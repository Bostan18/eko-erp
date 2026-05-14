import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


VERT_EKO   = "1A5C38"
VERT_CLAIR = "E8F5EE"

CAT_LABEL = {
    "salaire":        "Salaires",
    "materiel":       "Matériel",
    "carburant":      "Carburant",
    "sous_traitance": "Sous-traitance",
    "location":       "Location",
    "fourniture":     "Fournitures",
    "autre":          "Autre",
}


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


# ── Excel Facture ──────────────────────────────────────────────────────────────

def facture_excel(facture):
    wb = Workbook()
    ws = wb.active
    ws.title = "Facture"

    header_fill = PatternFill("solid", fgColor=VERT_EKO)
    bold        = Font(bold=True)
    white_bold  = Font(bold=True, color="FFFFFF")

    ws.merge_cells("A1:E1")
    ws["A1"] = f"FACTURE {facture.numero_local}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    infos = [
        ("Client",   facture.client.nom),
        ("Projet",   facture.projet.nom if facture.projet else "—"),
        ("Date",     str(facture.created_at.date())),
        ("Échéance", str(facture.date_echeance) if facture.date_echeance else "—"),
        ("Statut",   facture.get_statut_display()),
    ]
    for i, (label, val) in enumerate(infos, start=2):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = bold
        ws[f"B{i}"] = val
        ws.merge_cells(f"B{i}:E{i}")

    row = 8
    headers = ["Désignation", "Quantité", "Prix unitaire (F)", "Total HT (F)", "Total TTC (F)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = white_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    for ligne in facture.lignes.all():
        row += 1
        ws.cell(row=row, column=1, value=ligne.designation).border = _border()
        ws.cell(row=row, column=2, value=float(ligne.quantite)).border = _border()
        ws.cell(row=row, column=3, value=float(ligne.prix_unitaire)).border = _border()
        ws.cell(row=row, column=4, value=float(ligne.total_ht)).border = _border()
        ws.cell(row=row, column=5, value=float(ligne.montant_ttc)).border = _border()

    row += 2
    totaux = [
        ("Total HT",      float(facture.total_ht)),
        ("Total TVA",     float(facture.total_tva)),
        ("TOTAL TTC",     float(facture.total_ttc)),
        ("Payé",          float(facture.montant_paye)),
        ("Solde restant", float(facture.solde_restant)),
    ]
    for label, val in totaux:
        ws.cell(row=row, column=4, value=label).font = bold
        cell = ws.cell(row=row, column=5, value=val)
        cell.font = Font(bold=(label in ("TOTAL TTC", "Solde restant")))
        cell.number_format = "#,##0"
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── Excel Charges ──────────────────────────────────────────────────────────────

def charges_excel(charges, titre="Charges"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Charges"

    header_fill = PatternFill("solid", fgColor=VERT_EKO)
    white_bold  = Font(bold=True, color="FFFFFF")
    bold        = Font(bold=True)

    ws.merge_cells("A1:G1")
    ws["A1"] = titre
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Date", "Libellé", "Catégorie", "Montant (F)", "Projet", "Fournisseur", "Référence"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = white_bold
        cell.fill = PatternFill("solid", fgColor="2D7A50")
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    total = 0
    for i, c in enumerate(charges, start=3):
        ws.cell(row=i, column=1, value=str(c.date)).border = _border()
        ws.cell(row=i, column=2, value=c.libelle).border = _border()
        ws.cell(row=i, column=3, value=CAT_LABEL.get(c.categorie, c.categorie)).border = _border()
        cell = ws.cell(row=i, column=4, value=float(c.montant))
        cell.number_format = "#,##0"
        cell.border = _border()
        ws.cell(row=i, column=5, value=c.projet.nom if c.projet else "—").border = _border()
        ws.cell(row=i, column=6, value=c.fournisseur or "—").border = _border()
        ws.cell(row=i, column=7, value=c.reference or "—").border = _border()
        total += float(c.montant)

    row = len(list(charges)) + 3
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "TOTAL"
    ws[f"A{row}"].font = bold
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=VERT_CLAIR)
    cell = ws.cell(row=row, column=4, value=total)
    cell.font = bold
    cell.number_format = "#,##0"
    cell.fill = PatternFill("solid", fgColor=VERT_CLAIR)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
