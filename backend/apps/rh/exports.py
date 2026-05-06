import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VERT_EKO   = "1A5C38"
VERT_CLAIR = "E8F5EE"

MOIS_FR = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
           'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def paie_excel(employes, presences_par_employe, mois, annee):
    """
    Génère la feuille de paie mensuelle.
    employes : queryset Employe
    presences_par_employe : dict {employe_id: [PresenceJournaliere, ...]}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Paie {MOIS_FR[mois]} {annee}"

    header_fill = PatternFill("solid", fgColor=VERT_EKO)
    sub_fill    = PatternFill("solid", fgColor=VERT_CLAIR)
    white_bold  = Font(bold=True, color="FFFFFF")
    bold        = Font(bold=True)

    # Titre
    ws.merge_cells("A1:G1")
    ws["A1"] = f"FEUILLE DE PAIE — {MOIS_FR[mois].upper()} {annee}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # En-tête colonnes
    headers = ["Code", "Nom & Prénom", "Type", "Taux/Salaire (F)", "Jours présents", "Total à payer (F)", "Poste"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = white_bold
        cell.fill = PatternFill("solid", fgColor="2D7A50")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[2].height = 30

    TYPE_LABEL = {"cdi": "CDI", "journalier": "Journalier", "moo": "MOO"}

    total_global = 0
    row = 3
    for emp in employes:
        presences = presences_par_employe.get(emp.id, [])
        jours_presents = len([p for p in presences if p.present])

        if emp.type_contrat == "cdi":
            taux = float(emp.salaire_mensuel or 0)
            total = taux
        else:
            taux = float(emp.taux_journalier or 0)
            total = sum(float(p.montant_du) for p in presences if p.present)

        total_global += total

        ws.cell(row=row, column=1, value=emp.code).border = _border()
        ws.cell(row=row, column=2, value=f"{emp.nom} {emp.prenom}").border = _border()
        ws.cell(row=row, column=3, value=TYPE_LABEL.get(emp.type_contrat, emp.type_contrat)).border = _border()

        cell = ws.cell(row=row, column=4, value=taux)
        cell.number_format = "#,##0"
        cell.border = _border()

        ws.cell(row=row, column=5, value=jours_presents if emp.type_contrat != "cdi" else "—").border = _border()

        cell = ws.cell(row=row, column=6, value=total)
        cell.font = bold
        cell.number_format = "#,##0"
        cell.border = _border()

        ws.cell(row=row, column=7, value=emp.poste or "—").border = _border()
        row += 1

    # Ligne total
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "MASSE SALARIALE TOTALE"
    ws[f"A{row}"].font = bold
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=VERT_CLAIR)
    ws[f"A{row}"].alignment = Alignment(horizontal="right")

    cell = ws.cell(row=row, column=6, value=total_global)
    cell.font = Font(bold=True, size=11)
    cell.number_format = "#,##0"
    cell.fill = PatternFill("solid", fgColor=VERT_CLAIR)

    # Largeurs
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
